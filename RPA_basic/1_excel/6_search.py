from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    #番号、英語、数学
    if int(row[1].value) > 80:
        print(row[0].value, "番目の学生は優等生")


for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "英語":
            cell.value = "パソコン"

wb.save("sample_modified.xlsx")

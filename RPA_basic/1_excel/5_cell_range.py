from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

#1行ずつData入れる
ws.append(["番号", "英語", "数学"]) # A, B, C
for i in range(1, 11):
    ws.append([i, randint(0,100), randint(0, 100)])


# col_B = ws["B"] #英語のcolumnだけを呼んでくる

# #for cell in col_B:
#     #print(cell.value)
# col_range = ws["B:C"] #英語、数学のcolumn持ってくる
# # for cols in col_range:
# #     for cell in cols:
# #         print(cell.value)

# row_title = ws[1] # 1番目のrowだけを持ってくる
# # for cell in row_title:
# #     print(cell.value)

# row_range = ws[2:6] #2番目から６番目まで持ってくる
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end="")
#     print()


# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] #２番目の行から最後の行まで
# for rows in row_range:
#     for cell in rows:
#         xy = coordinate_from_string(cell.coordinate)
#         # print(xy, end="")
#         print(xy[0], end="")
#         print(xy[1], end=" ")
#     print()

# for column in tuple(ws.columns):
#     print(column[0].value)

for row in ws.iter_rows(min_row=1, max_row=5):
    print(row[2].value)

wb.save("sample.xlsx")
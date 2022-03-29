from openpyxl import Workbook
wb = Workbook()
ws = wb.active

#現在、作成した最終成績データを入力
ws.append(("学番", "出席", "Quiz1", "Quiz2", "中間テスト", "期末テスト", "プロジェクト"))

scores = [
(1, 10, 8, 5, 14, 26, 12),
(2, 7, 3, 7, 15, 24, 18),
(3, 9, 6, 8, 8, 12, 4),
(4, 7, 8, 7, 17, 21, 18),
(5, 7, 8, 7, 16, 25, 15),
(6, 3, 5, 8, 8, 17, 0),
(7, 4, 9, 10, 16, 27, 18),
(8, 6, 6, 15, 19, 17),
(9, 10, 10, 9, 19, 30, 19),
(10, 9, 8, 8, 20, 25, 20)
]

for s in scores:
    ws.append(s)

# 1.Quiz2の点数をすべて10に修正
for idx, cell in enumerate(ws["D"]):
    if idx == 0: # タイトルの場合は Skip
        continue
    cell.value = 10

# 2.H例に総点(SUM利用), I列に成績情報追加
ws["H1"] = "総点"
ws["I1"] = "成績情報"


# a b c d e f g h i
# 1 2 3 4 5 6 7 8 9
for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:]) - score[3] + 10 #総点データ
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)
    #SUM(B2:G2)
    #SUM(B3:G3)...
    
    # - 総点90点以上 A, 80点以上 B, 70点以上 C, その他 D
    grade = None #成績
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"

    # 3.出席が 5未満の学生は総点関係なく F
    if score[1] < 5:
        grade = "F"
    
    ws.cell(row=idx, column=9).value = grade #I列に成績情報

wb.save("scores.xlsx")
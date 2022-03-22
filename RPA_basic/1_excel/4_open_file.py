from openpyxl import load_workbook #fileをload
wb = load_workbook("sample.xlsx") #sample.xlsx fileでwbをload
ws = wb.active #活性化したSheet

# cell Data 呼んでくる

for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end="") #1, 2, 3, 4...
    print()

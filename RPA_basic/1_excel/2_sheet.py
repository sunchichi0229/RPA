from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() #新しいSheetを基本名で生成
ws.title = "MySheet" #Sheet名変更
ws.sheet_properties.tabColor = "ff66ff" #RPGの形態で色変更

ws1 = wb.create_sheet("YourSheet") #与えられた名前でSheet生成
ws2 = wb.create_sheet("NewSheet", 2) #2番目のindexにSheet生成

new_ws = wb["NewSheet"] #Dict形態でSheetに接近

print(wb.sheetnames) #すべてのSheet名を確認

# Sheet　コピー

new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
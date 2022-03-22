from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "SuSheet"

#　Aｎセルにｎの値を入力

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) #A1の情報を出力
print(ws["A1"].value) #A1の値を出力
print(ws["A10"]) #値がない時は'none'

# row = 1, 2, 3.....
# column = A(1), B(2), C(3), ....
print(ws.cell(column=1, row=1).value)

c = ws.cell(column=3, row=1, value=10)

from random import *

index = 1
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("sample.xlsx")